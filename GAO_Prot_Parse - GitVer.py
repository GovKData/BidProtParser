## Imports ##
import sys
reload(sys)
sys.setdefaultencoding('utf-8')
import codecs
import os
from os import listdir
import openpyxl
import string
import re
import random
import time



##############################
## READ THIS BEFORE STARTING
###########################

## 1 - Run in python, not python3

## 2 - This outputs both all protests and in five year
##    -- increments. As such, in the post-processing
##    -- stages, there are a number of large sections
##    -- which are functionally redundant. Be careful
##    -- modifying them. 

## 3 - See the index below, the "q-" suffix, paired
##    -- paired with the index number is meant to 
##    -- allow for a quick to ctrl+f and hop to 
##    -- the relevant section. 

## 4 - The larger the number of pound symbols 
##    -- surrounding headers, the more significant
##    -- the section break is

## 5(a) - Single pound symbols are used to comment out
##    -- pieces of code, to turn them "on" or "off"
##    -- e.g. there is a section where you can automate
##    -- the output of data or not, depending on
##    -- what you are working on with the program.
## 5(b) - Double pound symbols indicate commendary / notation
 

##############
## INDEX
############

## q-1) Published Merits Decision Handler
## q-2) Build Out Excel Sheet
## q-3) Post Processing
## q-4) File Output Management


######################################
####################################
###
##### SET-UP 
###
#################################
###############################

start_time = time.time()

## Enter the Input Folder Location, where all
## -- the files to be processed are
Inp_FN = '/home/owner/Desktop/Deprecated Prot Parser files/GAODir/ClnTxtFs/'

## Enter the Destination Folder Name Here:
Dest_Folder = '/home/owner/Desktop/FINAL PROT PARSER/Outputs/'

## Create Root File Name for the save files (the suffix that
## -- that will precede them at output
RFN = ('FinOut Full ')

## Lists to be used

Alph_F_Names = [] # Holds names for organizing

All_Ps_Doc = [] # All protests doc file
All_Ps_CSV_r = [] # All protests raw file, needed for handing before the 
                  ## extrapolations are added

All_Ps_CSV = [] # All protests csv file 
Sust_Ps_Doc = [] # Sustained protests doc file
Sust_Ps_CSV = [] # Sustained protests csv file

All_90_94_Doc = ['1990-1994']
All_95_99_Doc = ['1995-1999']
All_00_04_Doc = ['2000-2004']
All_05_09_Doc = ['2005-2009']
All_10_14_Doc = ['2010-2014']
All_15_19_Doc = ['2015-2020']
All_20_24_Doc = ['2020-2024']

Sust_90_94_Doc = ['1990-1994']
Sust_95_99_Doc = ['1995-1999']
Sust_00_04_Doc = ['2000-2004']
Sust_05_09_Doc = ['2005-2009']
Sust_10_14_Doc = ['2010-2014']
Sust_15_19_Doc = ['2015-2019']
Sust_20_24_Doc = ['2020-2024']

All_90_94_CSV_r = ['1990-1994']
All_95_99_CSV_r = ['1995-1999']
All_00_04_CSV_r = ['2000-2004']
All_05_09_CSV_r = ['2005-2009']
All_10_14_CSV_r = ['2010-2014']
All_15_19_CSV_r = ['2015-2019']
All_20_24_CSV_r = ['2020-2024']

Sust_90_94_CSV_r = ['1990-1994']
Sust_95_99_CSV_r = ['1995-1999']
Sust_00_04_CSV_r = ['2000-2004']
Sust_05_09_CSV_r = ['2005-2009']
Sust_10_14_CSV_r = ['2010-2014']
Sust_15_19_CSV_r = ['2015-2019']
Sust_20_24_CSV_r = ['2020-2024']

All_90_94_CSV = []
All_95_99_CSV = []
All_00_04_CSV = []
All_05_09_CSV = []
All_10_14_CSV = []
All_15_19_CSV = []
All_20_24_CSV = []

Sust_90_94_CSV = []
Sust_95_99_CSV = []
Sust_00_04_CSV = []
Sust_05_09_CSV = []
Sust_10_14_CSV = []
Sust_15_19_CSV = []
Sust_20_24_CSV = []

Assholes = [] # CSV file with all the problematic cases

##############################
## Create Macro Data Headers
########################

## This information will sit at the top of the spreadsheet 
## -- and collates the information in the rest of the sheet. 
## -- it uses a "xqx" prefix to "smuggle" the formula into excel
## -- so that the cell later performs the proper calculation. 

## This needs to be up here so that the program can retrieve the linebreak count
## -- variable for each line

BRK = '    <<|>>    '

## Macro Data Extrapolations, with commentary inset to populate 
## -- two cells over in the spreadsheet

MD_tot_prots = "xqx=SUM(D[s]:D[e])"
MD_tot_sol_protd = 'xqx=SUM(IF(1/COUNTIF(F[s]:F[e][XCLCOM]F[s]:F[e])=1[XCLCOM]1[XCLCOM]0))'
MD_avg_prot_fld = "xqx=AVERAGE(Y[s]:Y[e])"
MD_most_prot_p_Ktr = "xqx=MAX(Y[s]:Y[e])"
MD_avg_prot_p_solic = "xqx=AVERAGE(R[s]:R[e])"
MD_max_prot_p_solic = "xqx=LARGE(F[s]:F[e][XCLCOM]((MAX($F$[s]:$F$[e]))+1))"
MD_AVCA_pctg = 'xqx=(COUNTIF($L$[s]:$L$[e][XCLCOM] "AVCA"))'
MD_KVWD = 'xqx=COUNTIF($L$[s]:$L$[e][XCLCOM] "KVWD")'
MD_sust_no_rcmp = 'xqx=COUNTIF(M[s]:M[e][XCLCOM] "0")'
MD_sust_won_rcmp = 'xqx=COUNTIFS(M[s]:M[e][XCLCOM]"<=1"[XCLCOM]N[s]:N[e][XCLCOM]"1")'
MD_sust_lost_rcmp = 'xqx=(COUNTIFS(M[s]:M[e][XCLCOM]"<=1"[XCLCOM]N[s]:N[e][XCLCOM]"<>1"))'
MD_avg_flng2res = 'xqx=AVERAGE(U[s]:U[e])'

## Text to Insert into the top of the document for explanations

TL1 = "BINARY KEY"
TL2 = "Prot Rep: 0 = Pro Se; 1 = Attorney Assisted"
TL3 = "Recomp. Modified?: 0 = Recompetition K not Modified; 1 = Recompetition K was Modified; 2 = Sustained but no Recompetition; 3 = Protest not sustained"
TL4 = "Ultimate K Awardee: 0 = Original Awardee; 1 = Protester; 2 = 3rd Party; 3 = Protest not sustained"
TL5 = "Small Business or Disadvantaged Group: 0 = large business; 1 = Small Business or Disadvantaged Group"
TL6 = " --------"
TL7 = "PLEASE NOTE INSTRUCTIONS TO GET THE SHEET WORKING:"
TL8 = "To solve the issue of there being commas in excel "
TL9 = "formulas and commas also being used to separate "
TL10 = "cells; so commas must be 'smuggled' in. Use Ctrl+H"
TL11 = "to first replace '[xclcom]' with a regular comma;"
TL12 = "then replace 'xqx=' with an equals sign."

## This builds out the headers / macro data which will be
## -- at the top of the spreadsheets. It looks a lot more
## -- confusing in code form. Open the .csv after the program
## -- has run and it'll make more sense. 

MD_Extraps_r = ("MACRO DATA\nTotal Protests Filed:,%s,,%s\nTotal Solicitations Protested:,%s,,%s\nAvg. No. of Prots Filed per Ktr:,%s,,%s\nMost Prots per Ktr:,%s,,%s\nAvg. Prots per Solicitation:,%s,,%s\nMost Protests for a Solicitation:,%s,,%s\nAVCA Count:,%s,,%s\nKVWD Count:,%s,,%s\nSustained but no Recompetition:,%s,,%s\nSustained and Protester won Recompetition:,%s,,%s\nSustained and Protester Lost Recompetition:,%s,,%s\nAverage Days Between Prot. Filing and Resolution:,%s,,%s\n\n\n" % (MD_tot_prots, TL1, MD_tot_sol_protd, TL2, MD_avg_prot_fld, TL3, MD_most_prot_p_Ktr, TL4, MD_avg_prot_p_solic, TL5, MD_max_prot_p_solic, TL6, MD_AVCA_pctg, TL7, MD_KVWD, TL8, MD_sust_no_rcmp, TL9, MD_sust_won_rcmp, TL10, MD_sust_lost_rcmp, TL11, MD_avg_flng2res, TL12))

xcl_S = MD_Extraps_r.count('\n') + 2
    ## you need to add 2 because python counts from 0 but excel counts from 1,
    ## -- so that puts you one integer behind, and then the linebreaks
    ## -- start out taking you from line 1 to 2, so they are one integer ahead. 

MD_Extraps = MD_Extraps_r.replace('[s]',str(xcl_S)) 
    ## [e] must be replaced later. 

##############################################
############################################
###
##### PUBLISHED MERITS DECISION HANDLER (q-1)
##### -- if only doing new protest
##### -- management, skip to the
##### -- "Build out Excel Sheet"
##### -- This section is only for 
##### -- data extraction for merits
##### -- decisions that are already
##### -- published.
###
########################################
######################################

#######################
## Alphabetize Files
#################

for filename in os.listdir(Inp_FN):

#    print("""\n=================================================""")    
    
    Lwr_FN = filename.lower()

    Str_2_Srt = ("%sQXSPLTXQ%s") % (Lwr_FN, filename)    
    Alph_F_Names.append(Str_2_Srt)
 
Alph_F_Names.sort()


  ###########################################
############### MAIN FOR LOOP ################
  ##########################################

## This iterates through each case in a given folder.
for count_v, Long_FN in enumerate(Alph_F_Names):
    if count_v == 0:
        print("Starting Main For-Loop\n")

    #print("""\n=================================================\n%s""" % filename)    
    Loop_start_time = time.time() 
    Long_FN_Idx = Long_FN.find("QXSPLTXQ")
    filename = Long_FN[Long_FN_Idx + 8 : ]   
    
    FN_Idx_No = filename.find("B-") 
    FN_Clean_r = filename[ : FN_Idx_No]
    FN_Clean = FN_Clean_r.strip().replace(':','')

    local_file_unicode_1 = codecs.open(Inp_FN + filename, encoding='utf-8')

    SF1_r =  local_file_unicode_1.read().decode('utf8')

    SF1 = SF1_r.replace("\n1\n"," ").replace("\n2\n"," ").replace("\n3\n"," ").replace("\n4\n"," ").replace("\n5\n"," ")

    
    ####################
    ## Get Title Info
    ##############
    
    ## Find Date
    CDate_IdxS = filename.rfind("(")
    CDate_IdxE = filename.rfind(")")    
    CDate_r1 = (filename[ CDate_IdxS +1 : CDate_IdxE ])    
    CDate = CDate_r1.strip()
    C_D_len = len(CDate)
    CYear = CDate[-4 : ]
   
    ## Find Case No.    
    CNo_Idx_S = filename.find('B-')
    CNo_Idx_E = filename.find('(') 
    CNo_r = filename[CNo_Idx_S : CNo_Idx_E]
    CNo = CNo_r.rstrip()    

    ###################################
    ## Clean off Page Break Material
    ############################

#    Pg_Brk_E = (("(%s)\n") % (CYear))
#    Pg_Brk_E_len = len(Pg_Brk_E)        
#    Pg_Brk_Idx_S = SF1.find("\n")  
#    Pg_Brk_Idx_E = SF1.find(Pg_Brk_E, Pg_Brk_Idx_S)  
#   Pg_Brk_Str = SF1[Pg_Brk_Idx_S - 6 : Pg_Brk_Idx_E + Pg_Brk_E_len]
        
    SF1_pg_brk_strip = SF1.replace('\n',' ')    

#    if len(Pg_Brk_Str) >= 10:
#        SF1_Cln = SF1_pg_brk_strip.replace('**','*').replace('**','*').replace("*1 ","\n").replace("*2 ","\n")    
#    if len(Pg_Brk_Str) < 10:
#        SF1_Cln = SF1.replace('**','*').replace('**','*').replace("*1 ","\n").replace("*2 ","\n")    

    SF1_Cln = SF1_pg_brk_strip    

    ####################
    ## Main Text Body 
    ##############
     
    ## Isolate Body
    start_body = ('\nDIGEST\n')
    Dec_Flg = ('\nDECISION\n')
      
    WeFlg = ('We .* prote.*\.')
    ProtFlg = ('The protest .*\.')
    ClaimFlg = ('The claim .*\.')
    DismFlg = ('The dismissal .*\.')
    WeTrad = ('We ')
    PROTCAPS = ('PROTEST')
    NBCONCAPS = ('WILL NOT BE CONSIDERED')

    SF1_We_find = re.findall(WeFlg,SF1_Cln)
    SF1_Prot_find = re.findall(ProtFlg, SF1_Cln)
    SF1_Claim_find = re.findall(ClaimFlg, SF1_Cln)
    SF1_Dism_find = re.findall(DismFlg, SF1_Cln)
    SF1_WeBrd_findS = SF1_Cln.find(WeTrad)
    SF1_WeBrd_findE = SF1_Cln.find('.', SF1_WeBrd_findS)

    SF1_PROTCAPS_find = SF1_Cln.rfind(PROTCAPS)
    SF1_PROTCAPS_findS = SF1_Cln.rfind("\n", SF1_PROTCAPS_find)    
    SF1_PROTCAPS_findE = SF1_Cln.find(".", SF1_PROTCAPS_find)

    SF1_NBCONCAPS_find = SF1_Cln.rfind(NBCONCAPS)
    SF1_NBCONCAPS_findS = SF1_Cln.rfind("\n", SF1_NBCONCAPS_find)    
    SF1_NBCONCAPS_findE = SF1_Cln.find(".", SF1_NBCONCAPS_find)
    
    SF1_We_len = len(SF1_We_find)
    SF1_Prot_len = len(SF1_Prot_find)
    SF1_Claim_len = len(SF1_Claim_find)
    SF1_Dism_len = len(SF1_Dism_find)
#    print("QFR: %s,%s,%s,%s,%s" % (SF1_We_len, SF1_Prot_len, SF1_Claim_len, SF1_Dism_len, SF1_WeBrd_findS))

    ## For Loop Lists
    Outcome_list = []
    End_Idx_list = []

    if len(SF1_We_find) != 0:
        SF1_We_Idx_r = SF1.find(SF1_We_find[0])
        We_len = len(SF1_We_find[0])
        SF1_We_Idx = SF1_We_Idx_r + We_len
        Outcome_list.append(SF1_We_find[0])
        End_Idx_list.append(SF1_We_Idx)

    if len(SF1_Prot_find) != 0:
        SF1_Prot_Idx_r = SF1.find(SF1_Prot_find[0])
        Prot_len = len(SF1_Prot_find[0])
        SF1_Prot_Idx = SF1_Prot_Idx_r + Prot_len
        Outcome_list.append(SF1_Prot_find[0])
        End_Idx_list.append(SF1_Prot_Idx)

    if len(SF1_Claim_find) != 0:
        SF1_Claim_Idx_r = SF1.find(SF1_Claim_find[0])
        Claim_len = len(SF1_Claim_find[0])
        SF1_Claim_Idx = SF1_Claim_Idx_r + Claim_len
        Outcome_list.append(SF1_Claim_find[0])
        End_Idx_list.append(SF1_Claim_Idx)
    
    if len(SF1_Dism_find) != 0:
        SF1_Dism_Idx_r = SF1.find(SF1_Dism_find[0])
        Dism_len = len(SF1_Dism_find[0])
        SF1_Dism_Idx = SF1_Dism_Idx_r + Dism_len
        Outcome_list.append(SF1_Dism_find[0])
        End_Idx_list.append(SF1_Dism_Idx)

    if SF1_PROTCAPS_find != -1:
        Outcome_list.append(SF1[SF1_PROTCAPS_findS : SF1_PROTCAPS_findE])
        End_Idx_list.append(SF1_PROTCAPS_findE)
    
    if SF1_NBCONCAPS_find != -1:
        Outcome_list.append(SF1[SF1_NBCONCAPS_findS : SF1_NBCONCAPS_findE])
        End_Idx_list.append(SF1_NBCONCAPS_findE)

    if SF1_WeBrd_findS != -1:
        Outcome_list.append(SF1_Cln[SF1_WeBrd_findS : SF1_WeBrd_findE])
        End_Idx_list.append(SF1_WeBrd_findE)

    End_Idx_list.sort()    

    if len(Outcome_list) == 0:
        Outcome_list.append("NO OUTCOME")
    if len(End_Idx_list) == 0:
        End_Idx_list.append(1751)
  

    ##################
    ## Build Output
    ##############

    Outcome = Outcome_list[0]
    
    ## Creates the "Body Text" that will be output to the .txt files
    SF1_body_Sval = SF1_Cln.find(start_body)    
    SF1_body_Eval = End_Idx_list[0]
    body_text_r1 = SF1_Cln[SF1_body_Sval : SF1_body_Eval]

    body_text = body_text_r1.replace("\n"," ").replace(" DIGEST","DIGEST:").replace("DECISION","\n\nDECISION:").replace(' 1 ',' ').replace('\n1 ',' ')
    
    body_len = len(body_text)
    
#    if body_len <= 20:
#        print(FN_Clean)

    ## Check Decision Length

    Dec_len_Idx_S = body_text.find("DECISION")
    Dec_isol = body_text[Dec_len_Idx_S : ]
    Dec_len = len(Dec_isol)    
    BT_to_Dec = body_text[ : Dec_len_Idx_S + 10]
    Dec_Switch = "Decision omitted for brevity."
    BT_Dec_trim = "%s%s" % (BT_to_Dec, Dec_Switch)
    if Dec_len >= 1750:
        body_text = BT_Dec_trim
        


#########################################################
#######################################################
###
##### BUILD OUT EXCEL SHEET (q-2)
##### -- This part is where the new inputs would be 
##### -- drawn from. Once these tethers were established
##### -- everything covering the Published Merits
##### -- Decision Handler could be deleted. 
###
###################################################
#################################################
## Openpyxl is the best program for .xlsx files
## -- because pandas has removed support for that
## -- file type. Because Pandas is a more flexible
## -- and well known tool, it is still good to
## -- convert the files from Openpyxl to Pandas

#######################################
## Open files and convert to Pandas
####################################
## NOTE: There are examples below of relevant calls
## -- via Pandas and Openpyxl. Pandas is better for
## -- some things (e.g. ripping full rows of data
## -- and trickier calls) and Openpyxl has better
## -- iteration functions. What is included below
## -- is just an example of the tooling and will
## -- have to be overhauled depending on the
## -- data-states on relevant servers. Skip to the
## -- "Pull info from dataframes" section (the next
## -- header) for more specific information.

    Del_path = "C:/Users/Will Dawson/Desktop/Spring 2022/Delete1.xlsx"
    Xmp_path = "C:/Users/Will Dawson/Desktop/Personal/Example.xlsx"

    # To open the workbook in openpyxl, a workbook object is created
    Del_wb = openpyxl.load_workbook(Del_path)
    Xmp_wb = openpyxl.load_workbook(Xmp_path)
     
    # Get workbook active sheet object from the active attribute
    Del_sheet = Del_wb.active
    Xmp_sheet = Xmp_wb.active

    ## Convert to Pandas
    ## NOTE: In Openpyxl, you have to load the workbook
    ## -- THEN load the sheet, then convert that to
    ## -- Pandas for it to function.

    Del_pdDF = pd.DataFrame(Del_sheet.values)
    Xmp_pdDF = pd.DataFrame(Xmp_sheet.values)

    print(Del_pdDF.head(3))
    # Cell objects also have a row, column, and coordinate attributes that provide location information for the cell.
    ## to get the full column or full row, do the following

    ## - Find the max for each column and row

    Del_max_col = Del_sheet.max_column
    Xmp_max_col = Xmp_sheet.max_column

    Del_max_row = Del_sheet.max_row
    Xmp_max_row = Xmp_sheet.max_row

    ## - Write for loops to iterate through each

    #*print("Del row 1 all columns")

    for v_cell in range(1, Del_max_col +1):
            vDel_cell = Del_sheet.cell(row = 1, column = v_cell)
            #*print(vDel_cell.value)

    #*print("Xmp row 1 all columns")

    for v_cell in range(1, Xmp_max_col +1):
            vXmp_cell = Xmp_sheet.cell(row = 1, column = v_cell)
            #*print(vXmp_cell.value)

    #*print("Del Col 1 all Rows")

    for v_cell in range(1, Del_max_row +1):
            vDel_cell = Del_sheet.cell(row = v_cell, column = 1)
            #*print(vDel_cell.value)

    #*print("Xmp Col 1 all Rows")

    for v_cell in range(1, Xmp_max_row +1):
            vXmp_cell = Xmp_sheet.cell(row = v_cell, column = 1)
            #*print(vXmp_cell.value)

    #*print("All Rows, All Columns")

    for v_cell in range(1, Xmp_max_row +1):
            vXmp_cell = Xmp_sheet.cell(row = v_cell, column = v_cell)
            #*print(vXmp_cell.value)

    # Note: The first row or column integer is 1, not 0.
     
    # Cell object is created by using sheet object's cell() method.
    Del_cell = Del_sheet.cell(row = 1, column = 1)
    Xmp_cell = Xmp_sheet.cell(row = 1, column = 1)

    #*print(("Del = %s \nXmp = %s") % (Del_cell.value, Xmp_cell.value))

################################
## Pull info from dataframes
#############################
## Note: The information here is for the generation of
## -- example sheets. As such, much of the data is
## -- randomly generated for now and wil need to be
## -- pointed to at the time of implementation. The
## -- value of keeping the filler material until the
## -- actual values are implemented is that it allows
## -- the overall output to be experimented with and
## -- checked as a total system, and ensure that any
## -- modifications or data calls that "break" the
## -- system are detectable, as iterative checks can
## -- be run at each step in implementation relatively
## -- easily.

## Company Name
    Comp_Name = FN_Clean
    ## Could be pulled from any number of related files / docs
    ## -- when logging new protests.

## DUNS No. 
    DUNS_no = random.choice(range(3000,4500))
    ## This will be able to be pulled from a number of gov't databases

## Prot Rep.
    ## This is imperfect 
    Date_idxS = SF1.find("Date:")
    prot_rep_S = SF1.find("\n", Date_idxS)
    prot_rep_E = SF1.find("protester", prot_rep_S)
      #print(SF1)
    if prot_rep_E == -1:
        prot_rep_E = SF1.find("requester", prot_rep_S)

    if prot_rep_E == -1:
        prot_rep_E = SF1.find("requestor", prot_rep_S)
    
    prot_rep_str = SF1[prot_rep_S : prot_rep_E]
    prot_rep_out = "1"    
    if " Esq." not in prot_rep_str:
        if " Law" not in prot_rep_str:
            prot_rep_out = "0"
   
## Short Cite
    ## Will need to be modified to target more complex array when / if 
    ## -- Agency / COFC decisions are rolled into the dataset
    SC_idxS = SF1.find("File: ")
    SC_r = (SF1[SC_idxS + 6 : Date_idxS-1])    
    short_cite = SC_r.strip()

## Contract Number
    ## This will have to be sourced from a government database, but will
    ## -- be easy to find. 
    K_no = random.choice(range(1000,5000))

## Solicitation Issue Date
    ## This will have to be sourced from a government database, but will
    ## -- be easy to find.
    rand_SID = random.choice(range(40179,43829))
    solic_iss_d = rand_SID

## Protest Filing Date
    ## This will have to be sourced from a government database, but will
    ## -- be easy to find.
    rand_PFD = rand_SID + random.choice(range(5,50))
    prot_fil_d = rand_PFD

## Protest Resolution Date
    
    ## ACTUAL SOLUTION
    prot_resl_d = SF1[Date_idxS + 6 : prot_rep_S]

    ## SIMULATED SOLUTION
    prot_resl_d = rand_PFD + random.choice(range(10,90))

## Ultimate Contract Award Date
    ## This will have to be sourced from a government database, but will
    ## -- be easy to find.
    ult_k_d =  prot_resl_d + random.choice(range(1,30))

## Originating Agency
    ## This will have to be sourced from a government database, but will
    ## -- be easy to find.
    orig_ag = random.choice(["DMV","SPCA","EPA","NASA","DPW","DOD","FBI","DON","DOI","BLM","DGIF","DGAF","USM"])

## Protest Outcome
    ## NOTE: THE ORDER OF THESE THREE ARE REALLY IMPORTANT! DO NOT
    ## -- MODIFY WITHOUT CAUSE: Sometimes GAO decisions dismiss, deny,
    ## -- and sustain in part, this structure ensures that if any part
    ## -- is dismissed or sustained, it will be recorded. The deprication
    ## -- order is intended to ensure that the most 'valuable' outcome 
    ## -- is preserved.
    
    ## vv Actual code to use vv
    
    #if "denied" in Outcome:
    #    prot_outc = "DMD"

    #if "dismissed" in Outcome:
    #    prot_outc = "DPD"

    #if "sustain" in Outcome:
    #    prot_outc = "SMD"

    #if [conditional]:
        #prot_outc = "AVCA"

    #if [conditional]:
        #prot_outc = "KVW"
    
    ## ^^ Actual Code ^^

    ## Simulated Code

    outc_rand = random.choice(["DMD","SMD","DMD","AVCA","KVWD"])

    prot_outc = outc_rand
  
## Recompetition Modification
    ## This will have to be sourced from a government database, and will
    ## -- be a more difficult input to automate, but could be done quickly
    ## -- manually.
    
    # if prot_outc != "SMD":
        #recomp_mod = 3 
        # Saves a lot of bullshit with excel for extrapolation

    ## temp output to simulate
    rcmp_mod = random.choice([0,1,2,3])

## Ultimate Awardee Record
    ## This will have to be sourced from a government database, but will
    ## -- be easy to source and input.
    
    ## vv Rough code for implementation vv

    # init_awd_DUNS = get initial awardee DUNS No from DB
    # ult_awd_DUNS = get ulitmate awardee DUNS No from DB

    # if ult_awd_DUNS == init_awd_DUNS:
        #ult_awd_rec = 0      

    # if ult_awd_DUNS == DUNS_no:
        #ult_awd_rec = 1      

    # if ult_awd_DUNS != DUNS_no:
        #if uld_awd_DUNS != init_awd_DUNS:
            #ult_awd_rec = 2    

    # if prot_outc != "SMD":
        #ult_awd_rec = 3 
        # Saves a lot of bullshit with excel for extrapolation

    ## ^^ Rough code for implementation ^^    

    ## temp output to simulate
    ult_awd_rec = random.choice([0,1,2,3])
    
## Was Contractor a Small Business or Disadvantaged Group?
    ## This will have to be sourced from a government database, but will
    ## -- be easy to source and input.

    sb_dg = random.choice([0,0,0,1])

## Procurement Value
    ## This will have to be sourced from a government database, but will
    ## -- be easy to source and input.

    proc_val = random.choice(range(1000000,4000000))

    ###################
    ## Extrapolations
    ###############

    ## This information will be provided by the information previously
    ## -- extracted above, the information below generates the formula 
    ## -- for a single row which can then be quickly modified to fit a given sheet
    ## -- and then 'drag and dropped' through the rest. 

    ## Individual Protest Extrapolations

    prot_p_k = "xqx=COUNTIF($E$%s:$E$[e][XCLCOM]E[xcl_X])" % (xcl_S)
    iss2prot = "xqx=H[xcl_X]-G[xcl_X]"
    prot2dec = "xqx=I[xcl_X]-H[xcl_X]" 
    prot2UltK = "xqx=J[xcl_X]-H[xcl_X]"
    FY = "xqx=IF(MONTH(G[xcl_X])<10[XCLCOM]YEAR(G[xcl_X])[XCLCOM](YEAR(G[xcl_X]))+1)"
    F2KDec = "xqx=J[xcl_X]-H[xcl_X]"
    PrePostAwd = "Pre/PostAwd" #THIS IS TBD
    prot_p_ktr = "xqx=COUNTIF($B$%s:$B$[e][XCLCOM]B[xcl_X])" % (xcl_S)


    ## Format for output to Documents

    BRK = '    <<|>>    '
    
    Doc_out_r = """\n{%s}

Case No.:%s -- Date: %s

Outcome: %s

%s.\n\n""" % (FN_Clean, CNo, CDate, Outcome, body_text)
    
    ## Standardize Corporation Types and other terms
    Doc_out = Doc_out_r.replace(' Inc ',' Inc. ').replace(' Inc, ',' Inc., ').replace(' LLC ',' L.L.C. ').replace(' LLC, ',' L.L.C., ').replace(' 1 ',' ').replace('  ',' ').replace('..','.')

    ## Replace all in text commas with spaces, then replace all 
    ## -- the @ symbols with commas for output to CSV files
    CSV_out_r1 = ("%s@%s@%s@%s@%s@%s@%s@%s@%s@%s@%s@%s@%s@%s@%s@%s@%s@%s@%s@%s@%s@%s@%s@%s" % (Comp_Name, DUNS_no, prot_rep_out, short_cite, K_no, solic_iss_d, prot_fil_d, prot_resl_d, ult_k_d, orig_ag, prot_outc, rcmp_mod, ult_awd_rec, sb_dg, proc_val,BRK, prot_p_k, iss2prot, prot2dec, prot2UltK, FY, F2KDec, PrePostAwd, prot_p_ktr))
    CSV_out_r2 = CSV_out_r1.replace(',',' ').replace('\n',' ').replace("January", "01-").replace("February", "02-").replace("March", "03-").replace("April", "04-").replace("May", "05-").replace("June", "06-").replace("July", "07-").replace("August", "08-").replace("September", "09-").replace("October", "10-").replace("November", "11-").replace("December", "12-").replace('--','-').replace('  ',' ')
    CSV_outr_3 = CSV_out_r2.replace('@',',')
    ## The line-break you want has to be inserted separately down here, 
    ## -- so that the unwanted line-breaks can be scrubbed out above.
    CSV_out = ('%s\n') % (CSV_outr_3)
#    print(CSV_out)    
    ## This second one is for the Asshole CSV especially, which omits the
    ## -- linebreak, since the reasons must be appended to the same line
    AHCSV_out = ('%s') % (CSV_outr_3)

    ## I leave these to be commented on and off as needed for troubleshooting
#    print(CSV_out)
#    print(Doc_out)

    ## Various criteria that would catch problematic protests
    FL_Aholes = []
    if SF1_body_Eval == -1:
        FL_Aholes.append(",SF1_body_Eval is -1")
    if CNo == -1:
        FL_Aholes.append(",CNo val is -1")
    if CDate == -1: 
        FL_Aholes.append("CDate val is -1") 
    if Outcome == "NO OUTCOME":
        FL_Aholes.append(",Outcome val is -1")
    if body_len <= 20:
        FL_Aholes.append(",body_text val is off")
    if "DECISION" not in body_text:
        FL_Aholes.append(",'DECISION' not in body")            
    if len(End_Idx_list) == 0:
        FL_Aholes.append(",End_Idx_list val is 0")
    if "protest" not in Outcome:
        if "We deny" in Outcome:
            continue
        if "We dismiss" in Outcome:
            continue
        if "We grant" in Outcome:
            continue
        if "be reimbursed $" in Outcome:
            continue
        if "We recommend" in Outcome:
            continue
        if "We deny the reconsideration request" in Outcome:
            continue
        FL_Aholes.append(",'Protest' Not in Outcome")               

    Ahole_flags = ",".join(FL_Aholes)
    Ahole_out = "%s,%s" % (AHCSV_out, Ahole_flags)
    if len(FL_Aholes) > 0:
        Assholes.append(Ahole_out)
        CSV_out = CSV_out.replace("\n", Ahole_flags + "\n")


    ########################
    ## Add to Lists
    ## -- This cycles assigns the protest to the relevant lists for 
    ## -- sorting the protests in five year increments
    ############ 

    ## Add to Macro List
    All_Ps_Doc.append(Doc_out)
    All_Ps_CSV_r.append(CSV_out)

    ## Add protests that were upheld
    if "sustain" in Outcome or "grant" in Outcome:
        Sust_Ps_Doc.append(Doc_out)
        Sust_Ps_CSV.append(CSV_out)

    ## Sort by 5 Year Chunk
    if CYear >= '1990' and CYear <= '1994':
        All_90_94_Doc.append(Doc_out)
        All_90_94_CSV_r.append(CSV_out)
        if "sustain" in Outcome or "grant" in Outcome:
            Sust_90_94_Doc.append(Doc_out)
            Sust_90_94_CSV_r.append(CSV_out)

    if CYear >= '1995' and CYear <= '1999':
        All_95_99_Doc.append(Doc_out)
        All_95_99_CSV_r.append(CSV_out)
        if "sustain" in Outcome or "grant" in Outcome:
            Sust_95_99_Doc.append(Doc_out)
            Sust_95_99_CSV_r.append(CSV_out)

    if CYear >= '2000' and CYear <= '2004':
        All_00_04_Doc.append(Doc_out)
        All_00_04_CSV_r.append(CSV_out)
        if "sustain" in Outcome or "grant" in Outcome:
            Sust_00_04_Doc.append(Doc_out)
            Sust_00_04_CSV_r.append(CSV_out)

    if CYear >= '2005' and CYear <= '2009':
        All_05_09_Doc.append(Doc_out)
        All_05_09_CSV_r.append(CSV_out)
        if "sustain" in Outcome or "grant" in Outcome:
            Sust_05_09_Doc.append(Doc_out)
            Sust_05_09_CSV_r.append(CSV_out)

    if CYear >= '2010' and CYear <= '2014':
        All_10_14_Doc.append(Doc_out)
        All_10_14_CSV_r.append(CSV_out)
        if "sustain" in Outcome or "grant" in Outcome:
            Sust_10_14_Doc.append(Doc_out)
            Sust_10_14_CSV_r.append(CSV_out)

    if CYear >= '2015' and CYear <= '2019':
        All_15_19_Doc.append(Doc_out)
        All_15_19_CSV_r.append(CSV_out)
        if "sustain" in Outcome or "grant" in Outcome:
            Sust_15_19_Doc.append(Doc_out)
            Sust_15_19_CSV_r.append(CSV_out)

    if CYear >= '2020' and CYear <= '2025':
        All_20_24_Doc.append(Doc_out)
        All_20_24_CSV_r.append(CSV_out)
        if "sustain" in Outcome or "grant" in Outcome:
            Sust_20_24_Doc.append(Doc_out)
            Sust_20_24_CSV_r.append(CSV_out)

####################
## Status count
## -- This just runs a counter during processing to update the user on
## -- how far along the program is
###############

    if str(count_v) == '250':
        print("\n'It's not dumb if it works' \n    - Aristotle\n\n")

    if (str(count_v)[-3 : ]) == '500':                
        print("Time to %s: %s" % (count_v, round((time.time()-start_time),2))) 

    if (str(count_v)[-3 : ]) == '000':    
        print("Time to %s: %s" % (count_v,round((time.time()-start_time),2)))

#Use for expediting run during troubleshooting
#    if count_v > 500:
#        print("BREAKING AFTER %s PASSES") % count_v        
#        break

############################################
##########################################
###
##### END OF THE PROCESSING FOR-LOOP
###
#######################################
#####################################



############################################
##########################################
###
##### START OF POST-PROCESSING (q-3)
##### -- This section is creating the
##### -- .csv files for each five year
##### -- group of protests
###
#######################################
#####################################

###########################
## Post-processing Work
## -- This section is creating the .csv files for each five year group
########################

print("\nPost-Processing Counting")

for full_var_str in All_Ps_CSV_r:
    All_Jnd = "".join(All_Ps_CSV_r)
    ## Create the end value for the spreadsheet
    xcl_end = str(len(All_Ps_CSV_r)+xcl_S-1) 
    ## Create the current line value for the spreadsheet
    xcl_x_var_int = (All_Ps_CSV_r.index(full_var_str)) + xcl_S
    xcl_x_var = str(xcl_x_var_int)    
    ## Count the number of times that general B no. occurs    
    short_cite_idx = full_var_str.find(",B-")
    short_cite_pull = full_var_str[short_cite_idx + 1 : short_cite_idx + 9]
    short_cite_count = All_Jnd.count(short_cite_pull)
    ## Create the new line and append
    new_var_str_r = (("%s,%s%s") % (full_var_str[0 : short_cite_idx], short_cite_count, full_var_str[short_cite_idx:]))
    new_var_str = new_var_str_r.replace("[xcl_X]", xcl_x_var).replace("[e]", xcl_end)
    if All_Ps_CSV_r.index(full_var_str) == 0:
        MD_Xtrp_out = MD_Extraps.replace('[e]',xcl_end)
        header = "%sCOMPANY NAME, DUNS  NO., PROT REP'd,CITE COUNT,SHORT CITE, CONTRACT NO., SOLIC ISS DATE, PROT FILE DATE, PROT DEC DATE, ULT K AWD DATE, ORIGINATING AGENCY, PROTEST OUTC, WAS K RECOMP'd, ULT K REC'pnt, SB or DG?, PROC VAL,<<|>>, PROT PER K, ISS2PROT, PROT2DEC, PROT2ULTK, FY, PROTF2KDEC, PrePostAwd, PROT PER KTR\n" % MD_Xtrp_out        
        All_Ps_CSV.append(header)    
    All_Ps_CSV.append(new_var_str)

for full_var_str in All_90_94_CSV_r:
    if All_90_94_CSV_r.index(full_var_str) == 0:
        continue    
    All_Jnd = "".join(All_90_94_CSV_r)
    ## Create the end value for the spreadsheet
    xcl_end = str(len(All_90_94_CSV_r)+xcl_S) 
    ## Create the current line value for the spreadsheet
    xcl_x_var_int = (All_90_94_CSV_r.index(full_var_str)) + xcl_S
    xcl_x_var = str(xcl_x_var_int)    
    ## Count the number of times that general B no. occurs
    short_cite_idx = full_var_str.find(",B-")
    short_cite_pull = full_var_str[short_cite_idx + 1 : short_cite_idx + 9]
    short_cite_count = All_Jnd.count(short_cite_pull)
    new_var_str_r = (("%s,%s%s") % (full_var_str[0 : short_cite_idx], short_cite_count, full_var_str[short_cite_idx:]))
    new_var_str = new_var_str_r.replace("[xcl_X]", xcl_x_var).replace("[e]", xcl_end)
    if All_90_94_CSV_r.index(full_var_str) == 0:
        MD_Xtrp_out = MD_Extraps.replace('[e]',xcl_end)
        header = "%sCOMPANY NAME, DUNS  NO., PROT REP'd,CITE COUNT,SHORT CITE, CONTRACT NO., SOLIC ISS DATE, PROT FILE DATE, PROT DEC DATE, ULT K AWD DATE, ORIGINATING AGENCY, PROTEST OUTC, WAS K RECOMP'd, ULT K REC'pnt, SB or DG?, PROC VAL,<<|>>, PROT PER K, ISS2PROT, PROT2DEC, PROT2ULTK, FY, PROTF2KDEC, PrePostAwd, PROT PER KTR\n" % MD_Xtrp_out        
        All_90_94_CSV.append(header)         
    All_90_94_CSV.append(new_var_str)

for full_var_str in All_95_99_CSV_r:
    if All_95_99_CSV_r.index(full_var_str) == 0:
        continue    
    All_Jnd = "".join(All_95_99_CSV_r)
    ## Create the end value for the spreadsheet
    xcl_end = str(len(All_95_99_CSV_r)+xcl_S) 
    ## Create the current line value for the spreadsheet
    xcl_x_var_int = (All_95_99_CSV_r.index(full_var_str)) + xcl_S
    xcl_x_var = str(xcl_x_var_int)    
    ## Count the number of times that general B no. occurs
    short_cite_idx = full_var_str.find(",B-")
    short_cite_pull = full_var_str[short_cite_idx + 1 : short_cite_idx + 9]
    short_cite_count = All_Jnd.count(short_cite_pull)
    new_var_str_r = new_var_str_r.replace("[xcl_X]", xcl_x_var).replace("[e]", xcl_end)    
    new_var_str = (("%s,%s%s") % (full_var_str[0 : short_cite_idx], short_cite_count, full_var_str[short_cite_idx:]))
    if All_95_99_CSV_r.index(full_var_str) == 0:
        MD_Xtrp_out = MD_Extraps.replace('[e]',xcl_end)
        header = "%sCOMPANY NAME, DUNS  NO., PROT REP'd,CITE COUNT,SHORT CITE, CONTRACT NO., SOLIC ISS DATE, PROT FILE DATE, PROT DEC DATE, ULT K AWD DATE, ORIGINATING AGENCY, PROTEST OUTC, WAS K RECOMP'd, ULT K REC'pnt, SB or DG?, PROC VAL,<<|>>, PROT PER K, ISS2PROT, PROT2DEC, PROT2ULTK, FY, PROTF2KDEC, PrePostAwd, PROT PER KTR\n" % MD_Xtrp_out        
        All_95_99_CSV.append(header)     
    All_95_99_CSV.append(new_var_str)

for full_var_str in All_00_04_CSV_r:
    if All_00_04_CSV_r.index(full_var_str) == 0:
        continue    
    All_Jnd = "".join(All_00_04_CSV_r)
    ## Create the end value for the spreadsheet
    xcl_end = str(len(All_00_04_CSV_r)+xcl_S) 
    ## Create the current line value for the spreadsheet
    xcl_x_var_int = (All_00_04_CSV_r.index(full_var_str)) + xcl_S
    xcl_x_var = str(xcl_x_var_int)    
    ## Count the number of times that general B no. occurs
    short_cite_idx = full_var_str.find(",B-")
    short_cite_pull = full_var_str[short_cite_idx + 1 : short_cite_idx + 9]
    short_cite_count = All_Jnd.count(short_cite_pull)
    new_var_str_r = (("%s,%s%s") % (full_var_str[0 : short_cite_idx], short_cite_count, full_var_str[short_cite_idx:]))
    new_var_str = new_var_str_r.replace("[xcl_X]", xcl_x_var).replace("[e]", xcl_end)
    if All_00_04_CSV_r.index(full_var_str) == 0:
        MD_Xtrp_out = MD_Extraps.replace('[e]',xcl_end)
        header = "%sCOMPANY NAME, DUNS  NO., PROT REP'd,CITE COUNT,SHORT CITE, CONTRACT NO., SOLIC ISS DATE, PROT FILE DATE, PROT DEC DATE, ULT K AWD DATE, ORIGINATING AGENCY, PROTEST OUTC, WAS K RECOMP'd, ULT K REC'pnt, SB or DG?, PROC VAL,<<|>>, PROT PER K, ISS2PROT, PROT2DEC, PROT2ULTK, FY, PROTF2KDEC, PrePostAwd, PROT PER KTR\n" % MD_Xtrp_out        
        All_00_04_CSV.append(header)         
    All_00_04_CSV.append(new_var_str)

for full_var_str in All_05_09_CSV_r:
    if All_05_09_CSV_r.index(full_var_str) == 0:
        continue    
    All_Jnd = "".join(All_05_09_CSV_r)
    ## Create the end value for the spreadsheet
    xcl_end = str(len(All_05_09_CSV_r)+xcl_S) 
    ## Create the current line value for the spreadsheet
    xcl_x_var_int = (All_05_09_CSV_r.index(full_var_str)) + xcl_S
    xcl_x_var = str(xcl_x_var_int)    
    ## Count the number of times that general B no. occurs
    short_cite_idx = full_var_str.find(",B-")
    short_cite_pull = full_var_str[short_cite_idx + 1 : short_cite_idx + 9]
    short_cite_count = All_Jnd.count(short_cite_pull)
    new_var_str_r = (("%s,%s%s") % (full_var_str[0 : short_cite_idx], short_cite_count, full_var_str[short_cite_idx:]))
    new_var_str = new_var_str_r.replace("[xcl_X]", xcl_x_var).replace("[e]", xcl_end)    
    if All_05_09_CSV_r.index(full_var_str) == 0:
        MD_Xtrp_out = MD_Extraps.replace('[e]',xcl_end)
        header = "%sCOMPANY NAME, DUNS  NO., PROT REP'd,CITE COUNT,SHORT CITE, CONTRACT NO., SOLIC ISS DATE, PROT FILE DATE, PROT DEC DATE, ULT K AWD DATE, ORIGINATING AGENCY, PROTEST OUTC, WAS K RECOMP'd, ULT K REC'pnt, SB or DG?, PROC VAL,<<|>>, PROT PER K, ISS2PROT, PROT2DEC, PROT2ULTK, FY, PROTF2KDEC, PrePostAwd, PROT PER KTR\n" % MD_Xtrp_out        
        All_05_09_CSV.append(header)     
    All_05_09_CSV.append(new_var_str)

for full_var_str in All_10_14_CSV_r:
    if All_10_14_CSV_r.index(full_var_str) == 0:
        continue    
    All_Jnd = "".join(All_10_14_CSV_r)
    ## Create the end value for the spreadsheet
    xcl_end = str(len(All_10_14_CSV_r)+xcl_S) 
    ## Create the current line value for the spreadsheet
    xcl_x_var_int = (All_10_14_CSV_r.index(full_var_str)) + xcl_S
    xcl_x_var = str(xcl_x_var_int)    
    ## Count the number of times that general B no. occurs
    short_cite_idx = full_var_str.find(",B-")
    short_cite_pull = full_var_str[short_cite_idx + 1 : short_cite_idx + 9]
    short_cite_count = All_Jnd.count(short_cite_pull)
    new_var_str_r = (("%s,%s%s") % (full_var_str[0 : short_cite_idx], short_cite_count, full_var_str[short_cite_idx:]))
    new_var_str = new_var_str_r.replace("[xcl_X]", xcl_x_var).replace("[e]", xcl_end)  
    if All_10_14_CSV_r.index(full_var_str) == 0:
        MD_Xtrp_out = MD_Extraps.replace('[e]',xcl_end)
        header = "%sCOMPANY NAME, DUNS  NO., PROT REP'd,CITE COUNT,SHORT CITE, CONTRACT NO., SOLIC ISS DATE, PROT FILE DATE, PROT DEC DATE, ULT K AWD DATE, ORIGINATING AGENCY, PROTEST OUTC, WAS K RECOMP'd, ULT K REC'pnt, SB or DG?, PROC VAL,<<|>>, PROT PER K, ISS2PROT, PROT2DEC, PROT2ULTK, FY, PROTF2KDEC, PrePostAwd, PROT PER KTR\n" % MD_Xtrp_out        
        All_10_14_CSV.append(header)       
    All_10_14_CSV.append(new_var_str)

for full_var_str in All_15_19_CSV_r:
    if All_15_19_CSV_r.index(full_var_str) == 0:
        continue    
    All_Jnd = "".join(All_15_19_CSV_r)
    ## Create the end value for the spreadsheet
    xcl_end = str(len(All_15_19_CSV_r)+xcl_S) 
    ## Create the current line value for the spreadsheet
    xcl_x_var_int = (All_15_19_CSV_r.index(full_var_str)) + xcl_S
    xcl_x_var = str(xcl_x_var_int)    
    ## Count the number of times that general B no. occurs
    short_cite_idx = full_var_str.find(",B-")
    short_cite_pull = full_var_str[short_cite_idx + 1 : short_cite_idx + 9]
    short_cite_count = All_Jnd.count(short_cite_pull)
    new_var_str_r = (("%s,%s%s") % (full_var_str[0 : short_cite_idx], short_cite_count, full_var_str[short_cite_idx:]))
    new_var_str = new_var_str_r.replace("[xcl_X]", xcl_x_var).replace("[e]", xcl_end)    
    if All_15_19_CSV_r.index(full_var_str) == 0:
        MD_Xtrp_out = MD_Extraps.replace('[e]',xcl_end)
        header = "%sCOMPANY NAME, DUNS  NO., PROT REP'd,CITE COUNT,SHORT CITE, CONTRACT NO., SOLIC ISS DATE, PROT FILE DATE, PROT DEC DATE, ULT K AWD DATE, ORIGINATING AGENCY, PROTEST OUTC, WAS K RECOMP'd, ULT K REC'pnt, SB or DG?, PROC VAL,<<|>>, PROT PER K, ISS2PROT, PROT2DEC, PROT2ULTK, FY, PROTF2KDEC, PrePostAwd, PROT PER KTR\n" % MD_Xtrp_out        
        All_15_19_CSV.append(header)    
    All_15_19_CSV.append(new_var_str)

for full_var_str in All_20_24_CSV_r:
    if All_20_24_CSV_r.index(full_var_str) == 0:
        continue    
    All_Jnd = "".join(All_20_24_CSV_r)
    ## Create the end value for the spreadsheet
    xcl_end = str(len(All_20_24_CSV_r)+xcl_S) 
    ## Create the current line value for the spreadsheet
    xcl_x_var_int = (All_20_24_CSV_r.index(full_var_str)) + xcl_S
    xcl_x_var = str(xcl_x_var_int)    
    ## Count the number of times that general B no. occurs
    short_cite_idx = full_var_str.find(",B-")
    short_cite_pull = full_var_str[short_cite_idx + 1 : short_cite_idx + 9]
    short_cite_count = All_Jnd.count(short_cite_pull)
    new_var_str_r = (("%s,%s%s") % (full_var_str[0 : short_cite_idx], short_cite_count, full_var_str[short_cite_idx:]))
    new_var_str = new_var_str_r.replace("[xcl_X]", xcl_x_var).replace("[e]", xcl_end)   
    if All_20_24_CSV_r.index(full_var_str) == 0:
        MD_Xtrp_out = MD_Extraps.replace('[e]',xcl_end)
        header = "%sCOMPANY NAME, DUNS  NO., PROT REP'd,CITE COUNT,SHORT CITE, CONTRACT NO., SOLIC ISS DATE, PROT FILE DATE, PROT DEC DATE, ULT K AWD DATE, ORIGINATING AGENCY, PROTEST OUTC, WAS K RECOMP'd, ULT K REC'pnt, SB or DG?, PROC VAL,<<|>>, PROT PER K, ISS2PROT, PROT2DEC, PROT2ULTK, FY, PROTF2KDEC, PrePostAwd, PROT PER KTR\n" % MD_Xtrp_out        
        All_20_24_CSV.append(header)     
    All_20_24_CSV.append(new_var_str)

## Join all items in the lists together

print("\nJoining Lists for Output")

AP_Doc = "".join(All_Ps_Doc)
AP_CSV = "".join(All_Ps_CSV)

AP_90_94_Doc = "".join(All_90_94_Doc)
AP_90_94_CSV = "".join(All_90_94_CSV)

AP_95_99_Doc = "".join(All_95_99_Doc)
AP_95_99_CSV = "".join(All_95_99_CSV)

AP_00_04_Doc = "".join(All_00_04_Doc)
AP_00_04_CSV = "".join(All_00_04_CSV)

AP_05_09_Doc = "".join(All_05_09_Doc)
AP_05_09_CSV = "".join(All_05_09_CSV)

AP_10_14_Doc = "".join(All_10_14_Doc)
AP_10_14_CSV = "".join(All_10_14_CSV)

AP_15_19_Doc = "".join(All_15_19_Doc)
AP_15_19_CSV = "".join(All_15_19_CSV)

AP_20_24_Doc = "".join(All_20_24_Doc)
AP_20_24_CSV = "".join(All_20_24_CSV)

SP_Doc = "".join(Sust_Ps_Doc)
SP_CSV = "".join(Sust_Ps_CSV)

SP_90_94_Doc = "".join(All_90_94_Doc)
SP_90_94_CSV = "".join(All_90_94_CSV)

SP_95_99_Doc = "".join(All_95_99_Doc)
SP_95_99_CSV = "".join(All_95_99_CSV)

SP_00_04_Doc = "".join(Sust_00_04_Doc)
SP_00_04_CSV = "".join(Sust_00_04_CSV)

SP_05_09_Doc = "".join(Sust_05_09_Doc)
SP_05_09_CSV = "".join(Sust_05_09_CSV)

SP_10_14_Doc = "".join(Sust_10_14_Doc)
SP_10_14_CSV = "".join(Sust_10_14_CSV)

SP_15_19_Doc = "".join(Sust_15_19_Doc)
SP_15_19_CSV = "".join(Sust_15_19_CSV)

SP_20_24_Doc = "".join(Sust_20_24_Doc)
SP_20_24_CSV = "".join(Sust_20_24_CSV)


Ass_CSV = "\n".join(Assholes)

## All Protests Document Directory Locations
AP_Doc_path = (Dest_Folder + RFN + 'All Protests Doc.txt')

AP_90_94_Doc_path = (Dest_Folder + RFN + '90-94 All Protests Doc.txt')
AP_95_99_Doc_path = (Dest_Folder + RFN + '95-99 All Protests Doc.txt')
AP_00_04_Doc_path = (Dest_Folder + RFN + '00-04 All Protests Doc.txt')
AP_05_09_Doc_path = (Dest_Folder + RFN + '05-09 All Protests Doc.txt')
AP_10_14_Doc_path = (Dest_Folder + RFN + '10-14 All Protests Doc.txt')
AP_15_19_Doc_path = (Dest_Folder + RFN + '15-19 All Protests Doc.txt')
AP_20_24_Doc_path = (Dest_Folder + RFN + '20-24 All Protests Doc.txt')

## All Protests CSV Directory Locations
AP_CSV_path = (Dest_Folder + RFN + 'All Protests CSV.csv')

AP_90_94_CSV_path = (Dest_Folder + RFN + '90-94 All Protests CSV.csv')
AP_95_99_CSV_path = (Dest_Folder + RFN + '95-99 All Protests CSV.csv')
AP_00_04_CSV_path = (Dest_Folder + RFN + '00-04 All Protests CSV.csv')
AP_05_09_CSV_path = (Dest_Folder + RFN + '05-09 All Protests CSV.csv')
AP_10_14_CSV_path = (Dest_Folder + RFN + '10-14 All Protests CSV.csv')
AP_15_19_CSV_path = (Dest_Folder + RFN + '15-19 All Protests CSV.csv')
AP_20_24_CSV_path = (Dest_Folder + RFN + '20-24 All Protests CSV.csv')

## All Sustained Protest Document Diectory Locations
SP_Doc_path = (Dest_Folder + RFN + 'Sust Protests Doc.txt')

SP_90_94_Doc_path = (Dest_Folder + RFN + '90-94 Sust Protests Doc.txt')
SP_95_99_Doc_path = (Dest_Folder + RFN + '95-99 Sust Protests Doc.txt')
SP_00_04_Doc_path = (Dest_Folder + RFN + '00-04 Sust Protests Doc.txt')
SP_05_09_Doc_path = (Dest_Folder + RFN + '05-09 Sust Protests Doc.txt')
SP_10_14_Doc_path = (Dest_Folder + RFN + '10-14 Sust Protests Doc.txt')
SP_15_19_Doc_path = (Dest_Folder + RFN + '15-19 Sust Protests Doc.txt')
SP_20_24_Doc_path = (Dest_Folder + RFN + '20-24 Sust Protests Doc.txt')

## All Sustained Protest CSV Diectory Locations
SP_CSV_path = (Dest_Folder + RFN + 'Sust Protests CSV.csv')

SP_90_94_CSV_path = (Dest_Folder + RFN + '90-94 Sust Protests CSV.csv')
SP_95_99_CSV_path = (Dest_Folder + RFN + '95-99 Sust Protests CSV.csv')
SP_00_04_CSV_path = (Dest_Folder + RFN + '00-04 Sust Protests CSV.csv')
SP_05_09_CSV_path = (Dest_Folder + RFN + '05-09 Sust Protests CSV.csv')
SP_10_14_CSV_path = (Dest_Folder + RFN + '10-14 Sust Protests CSV.csv')
SP_15_19_CSV_path = (Dest_Folder + RFN + '15-19 Sust Protests CSV.csv')
SP_20_24_CSV_path = (Dest_Folder + RFN + '20-24 Sust Protests CSV.csv')

Ass_CSV_path = (Dest_Folder + RFN + 'Protests Assholes.csv')

####################################################
##################################################
###
##### END OF POST PROCESSING
###
################################################
#############################################




####################################################
##################################################
###
##### START OF FILE OUTPUT MANAGEMENT (q-4)
##### -- This section is divided into outputs
##### -- for the documents and .csv files
###
################################################
#############################################


####################################
## Decide whether to output files
###############################

## Comment one of these next two lines out to turn the output on or off. 

#Create_Files_r = 'n'
Create_Files_r = raw_input("\nWould you like to create output files?\n(Y or N): ")

Create_Files = Create_Files_r.lower()

#########################
## File Output Handler
###################



if Create_Files == 'y':
    
    print("""

  ++++++++++++++++++++
+++ OUTPUTTING FILES +++
  ++++++++++++++++++++
""")

    
    ##################################################
    ################################################
    ### Doc Files
    #############################################
    ###########################################


    ## All Protests Doc
    with codecs.open (AP_Doc_path, 'w', 'utf-8') as fd:

        ## This inserts a custom header that can include any notes you want to leave for yourself during final formatting        
        fd.write("""Steps to bulk format: 
    1) Ctrl + H and type in '(\{)(*)(\})' to the Find bar and '/1/2/3' [But do the escape slashes, not fwd slashes] into the Replace bar. Then go to More Options, click 'Use wildcards' then below, click 'Format' and make whatever changes you want

    2) ^p replace the breaks you want with special characters, so ^p^p^p with QDBLQ for the double spaces you want to keep, then ^p^p with QSNGLQ for the single breaks you want, then clean all the bad \n breaks with ^p replace with null, then backfill QSNGLEQ and QDBLQ

+++++ DELETE ABOVE HERE BEFORE SHARING +++++++

    Note: This list is alphabetized. It is intended to be used in conjunction with a spread sheet listing all of these cases and their outcomes. The thought is that the spread sheet acts as an index to quickly identify all the sustained protests.\n\n""")     
        
        fd.write(AP_Doc)

        fd.close()  

    ## Sustained Protests Doc
    with codecs.open (SP_Doc_path, 'w', 'utf-8') as fd:

        ## This inserts a custom header that can include any notes you want to leave for yourself during final formatting         
        fd.write("""Steps to bulk format: 
    1) Ctrl + H and type in '(\{)(*)(\})' to the Find bar and '/1/2/3' [But do the escape slashes, not fwd slashes] into the Replace bar. Then go to More Options, click 'Use wildcards' then below, click 'Format' and make whatever changes you want

    2) ^p replace the breaks you want with special characters, so ^p^p^p with QDBLQ for the double spaces you want to keep, then ^p^p with QSNGLQ for the single breaks you want, then clean all the bad \n breaks with ^p replace with null, then backfill QSNGLEQ and QDBLQ

+++++ DELETE ABOVE HERE BEFORE SHARING +++++++

    Note: This list is alphabetized. It is intended to be used in conjunction with a spread sheet listing all of these cases and their outcomes. The thought is that the spread sheet acts as an index to quickly identify all the sustained protests.\n\n""")     
        
        fd.write(SP_Doc)
        
        fd.close()

## 1990 - 1994 
    ## All Protests Doc
    with codecs.open (AP_90_94_Doc_path, 'w', 'utf-8') as fd:
        ## Write in the column headers you want to use
        fd.write("""Note: This list is alphabetized. It is intended to be used in conjunction with a spread sheet listing all of these cases and their outcomes. The thought is that the spread sheet acts as an index to quickly identify all the sustained protests.\n\n""")     

        fd.write(AP_90_94_Doc)

        fd.close()  

    ## Sustained Protests Doc
    with codecs.open (SP_90_94_Doc_path, 'w', 'utf-8') as fd:
        ## Write in the column headers you want to use
        fd.write("""Note: This list is alphabetized. It is intended to be used in conjunction with a spread sheet listing all of these cases and their outcomes. The thought is that the spread sheet acts as an index to quickly identify all the sustained protests.\n\n""")      

        fd.write(SP_90_94_Doc)

        fd.close()  

## 1995 - 1999 
    ## All Protests Doc
    with codecs.open (AP_95_99_Doc_path, 'w', 'utf-8') as fd:
        ## Write in the column headers you want to use
        fd.write("""Note: This list is alphabetized. It is intended to be used in conjunction with a spread sheet listing all of these cases and their outcomes. The thought is that the spread sheet acts as an index to quickly identify all the sustained protests.\n\n""")     

        fd.write(AP_95_99_Doc)

        fd.close()  

    ## Sustained Protests Doc
    with codecs.open (SP_95_99_Doc_path, 'w', 'utf-8') as fd:
        ## Write in the column headers you want to use
        fd.write("""Note: This list is alphabetized. It is intended to be used in conjunction with a spread sheet listing all of these cases and their outcomes. The thought is that the spread sheet acts as an index to quickly identify all the sustained protests.\n\n""")      

        fd.write(SP_95_99_Doc)

        fd.close()  

## 2000 - 2004 
    ## All Protests Doc
    with codecs.open (AP_00_04_Doc_path, 'w', 'utf-8') as fd:
        ## Write in the column headers you want to use
        fd.write("""Note: This list is alphabetized. It is intended to be used in conjunction with a spread sheet listing all of these cases and their outcomes. The thought is that the spread sheet acts as an index to quickly identify all the sustained protests.\n\n""")     

        fd.write(AP_00_04_Doc)

        fd.close()  

    ## Sustained Protests Doc
    with codecs.open (SP_00_04_Doc_path, 'w', 'utf-8') as fd:
        ## Write in the column headers you want to use
        fd.write("""Note: This list is alphabetized. It is intended to be used in conjunction with a spread sheet listing all of these cases and their outcomes. The thought is that the spread sheet acts as an index to quickly identify all the sustained protests.\n\n""")      

        fd.write(SP_00_04_Doc)

        fd.close()  

## 2005 - 2009 
    ## All Protests Doc
    with codecs.open (AP_05_09_Doc_path, 'w', 'utf-8') as fd:
        ## Write in the column headers you want to use
        fd.write("""Note: This list is alphabetized. It is intended to be used in conjunction with a spread sheet listing all of these cases and their outcomes. The thought is that the spread sheet acts as an index to quickly identify all the sustained protests.\n\n""")     

        fd.write(AP_05_09_Doc)

        fd.close()  

    ## Sustained Protests Doc
    with codecs.open (SP_05_09_Doc_path, 'w', 'utf-8') as fd:
        ## Write in the column headers you want to use
        fd.write("""Note: This list is alphabetized. It is intended to be used in conjunction with a spread sheet listing all of these cases and their outcomes. The thought is that the spread sheet acts as an index to quickly identify all the sustained protests.\n\n""")      

        fd.write(SP_05_09_Doc)

        fd.close()  

## 2010 - 2014 
    ## All Protests Doc
    with codecs.open (AP_10_14_Doc_path, 'w', 'utf-8') as fd:
        ## Write in the column headers you want to use
        fd.write("""Note: This list is alphabetized. It is intended to be used in conjunction with a spread sheet listing all of these cases and their outcomes. The thought is that the spread sheet acts as an index to quickly identify all the sustained protests.\n\n""")     

        fd.write(AP_10_14_Doc)

        fd.close()  

    ## Sustained Protests Doc
    with codecs.open (SP_10_14_Doc_path, 'w', 'utf-8') as fd:
        ## Write in the column headers you want to use
        fd.write("""Note: This list is alphabetized. It is intended to be used in conjunction with a spread sheet listing all of these cases and their outcomes. The thought is that the spread sheet acts as an index to quickly identify all the sustained protests.\n\n""")      

        fd.write(SP_10_14_Doc)

        fd.close()  

## 2015 - 2019 
    ## All Protests Doc
    with codecs.open (AP_15_19_Doc_path, 'w', 'utf-8') as fd:
        ## Write in the column headers you want to use
        fd.write("""Note: This list is alphabetized. It is intended to be used in conjunction with a spread sheet listing all of these cases and their outcomes. The thought is that the spread sheet acts as an index to quickly identify all the sustained protests.\n\n""")     

        fd.write(AP_15_19_Doc)

        fd.close()  

    ## Sustained Protests Doc
    with codecs.open (SP_15_19_Doc_path, 'w', 'utf-8') as fd:
        ## Write in the column headers you want to use
        fd.write("""Note: This list is alphabetized. It is intended to be used in conjunction with a spread sheet listing all of these cases and their outcomes. The thought is that the spread sheet acts as an index to quickly identify all the sustained protests.\n\n""")      

        fd.write(SP_15_19_Doc)

        fd.close()  

## 2020 - 2024 
    ## All Protests Doc
    with codecs.open (AP_20_24_Doc_path, 'w', 'utf-8') as fd:
        ## Write in the column headers you want to use
        fd.write("""Note: This list is alphabetized. It is intended to be used in conjunction with a spread sheet listing all of these cases and their outcomes. The thought is that the spread sheet acts as an index to quickly identify all the sustained protests.\n\n""")     

        fd.write(AP_20_24_Doc)

        fd.close()  

    ## Sustained Protests Doc
    with codecs.open (SP_20_24_Doc_path, 'w', 'utf-8') as fd:
        ## Write in the column headers you want to use
        fd.write("""Note: This list is alphabetized. It is intended to be used in conjunction with a spread sheet listing all of these cases and their outcomes. The thought is that the spread sheet acts as an index to quickly identify all the sustained protests.\n\n""")      

        fd.write(SP_20_24_Doc)

        fd.close()  




    ##################################################
    ################################################
    ### CSV Files
    #############################################
    ###########################################
   

 
    ## All Protests CSV
    with codecs.open (AP_CSV_path, 'w', 'utf-8') as fd:
        ## Write in the column headers you want to use

        fd.write(AP_CSV)

        fd.close()  

    ## Sustained Protests CSV
    with codecs.open (SP_CSV_path, 'w', 'utf-8') as fd:
        ## Write in the column headers you want to use
               
        fd.write("COMPANY NAME, DUNS  NO.,PROT REP'd,PROTS_CONT'd,SHORT CITE,CONTRACT NO.,SOLIC ISS DATE,PROT FILE DATE,PROT DEC DATE,ULT K AWD DATE,ORIGINATING AGENCY,PROTEST OUTC,WAS K RECOMP'd,ULT K REC'pnt,SB or DG?,PROC VAL,<<|>>,PROT PER K,ISS2PROT,PROT2DEC,PROT2ULTK,FY,PROTF2KDEC,PrePostAwd,PROT PER KTR\n")      

        fd.write(SP_CSV)

        fd.close()  

## 1990 - 1994 
    ## All Protests CSV
    with codecs.open (AP_90_94_CSV_path, 'w', 'utf-8') as fd:
        ## Write in the column headers you want to use
 
        fd.write("COMPANY NAME, DUNS  NO.,PROT REP'd,PROTS_CONT'd,SHORT CITE,CONTRACT NO.,SOLIC ISS DATE,PROT FILE DATE,PROT DEC DATE,ULT K AWD DATE,ORIGINATING AGENCY,PROTEST OUTC,WAS K RECOMP'd,ULT K REC'pnt,SB or DG?,PROC VAL,<<|>>,PROT PER K,ISS2PROT,PROT2DEC,PROT2ULTK,FY,PROTF2KDEC,PrePostAwd,PROT PER KTR\n")     

        fd.write(AP_90_94_CSV)

        fd.close()  

    ## Sustained Protests CSV
    with codecs.open (SP_90_94_CSV_path, 'w', 'utf-8') as fd:
        ## Write in the column headers you want to use
 
        fd.write("COMPANY NAME, DUNS  NO.,PROT REP'd,PROTS_CONT'd,SHORT CITE,CONTRACT NO.,SOLIC ISS DATE,PROT FILE DATE,PROT DEC DATE,ULT K AWD DATE,ORIGINATING AGENCY,PROTEST OUTC,WAS K RECOMP'd,ULT K REC'pnt,SB or DG?,PROC VAL,<<|>>,PROT PER K,ISS2PROT,PROT2DEC,PROT2ULTK,FY,PROTF2KDEC,PrePostAwd,PROT PER KTR\n")      

        fd.write(SP_90_94_CSV)

        fd.close()  

## 1995 - 1999 
    ## All Protests CSV
    with codecs.open (AP_95_99_CSV_path, 'w', 'utf-8') as fd:
        ## Write in the column headers you want to use
 
        fd.write("COMPANY NAME, DUNS  NO.,PROT REP'd,PROTS_CONT'd,SHORT CITE,CONTRACT NO.,SOLIC ISS DATE,PROT FILE DATE,PROT DEC DATE,ULT K AWD DATE,ORIGINATING AGENCY,PROTEST OUTC,WAS K RECOMP'd,ULT K REC'pnt,SB or DG?,PROC VAL,<<|>>,PROT PER K,ISS2PROT,PROT2DEC,PROT2ULTK,FY,PROTF2KDEC,PrePostAwd,PROT PER KTR\n")     

        fd.write(AP_95_99_CSV)

        fd.close()  

    ## Sustained Protests CSV
    with codecs.open (SP_95_99_CSV_path, 'w', 'utf-8') as fd:
        ## Write in the column headers you want to use
 
        fd.write("COMPANY NAME, DUNS  NO.,PROT REP'd,PROTS_CONT'd,SHORT CITE,CONTRACT NO.,SOLIC ISS DATE,PROT FILE DATE,PROT DEC DATE,ULT K AWD DATE,ORIGINATING AGENCY,PROTEST OUTC,WAS K RECOMP'd,ULT K REC'pnt,SB or DG?,PROC VAL,<<|>>,PROT PER K,ISS2PROT,PROT2DEC,PROT2ULTK,FY,PROTF2KDEC,PrePostAwd,PROT PER KTR\n")      

        fd.write(SP_95_99_CSV)

        fd.close()  

## 2000 - 2004 
    ## All Protests CSV
    with codecs.open (AP_00_04_CSV_path, 'w', 'utf-8') as fd:
        ## Write in the column headers you want to use
 
        fd.write("COMPANY NAME, DUNS  NO.,PROT REP'd,PROTS_CONT'd,SHORT CITE,CONTRACT NO.,SOLIC ISS DATE,PROT FILE DATE,PROT DEC DATE,ULT K AWD DATE,ORIGINATING AGENCY,PROTEST OUTC,WAS K RECOMP'd,ULT K REC'pnt,SB or DG?,PROC VAL,<<|>>,PROT PER K,ISS2PROT,PROT2DEC,PROT2ULTK,FY,PROTF2KDEC,PrePostAwd,PROT PER KTR\n")     

        fd.write(AP_00_04_CSV)

        fd.close()  

    ## Sustained Protests CSV
    with codecs.open (SP_00_04_CSV_path, 'w', 'utf-8') as fd:
        ## Write in the column headers you want to use
 
        fd.write("COMPANY NAME, DUNS  NO.,PROT REP'd,PROTS_CONT'd,SHORT CITE,CONTRACT NO.,SOLIC ISS DATE,PROT FILE DATE,PROT DEC DATE,ULT K AWD DATE,ORIGINATING AGENCY,PROTEST OUTC,WAS K RECOMP'd,ULT K REC'pnt,SB or DG?,PROC VAL,<<|>>,PROT PER K,ISS2PROT,PROT2DEC,PROT2ULTK,FY,PROTF2KDEC,PrePostAwd,PROT PER KTR\n")      

        fd.write(SP_00_04_CSV)

        fd.close()  

## 2005 - 2009 
    ## All Protests CSV
    with codecs.open (AP_05_09_CSV_path, 'w', 'utf-8') as fd:
        ## Write in the column headers you want to use
 
        fd.write("COMPANY NAME, DUNS  NO.,PROT REP'd,PROTS_CONT'd,SHORT CITE,CONTRACT NO.,SOLIC ISS DATE,PROT FILE DATE,PROT DEC DATE,ULT K AWD DATE,ORIGINATING AGENCY,PROTEST OUTC,WAS K RECOMP'd,ULT K REC'pnt,SB or DG?,PROC VAL,<<|>>,PROT PER K,ISS2PROT,PROT2DEC,PROT2ULTK,FY,PROTF2KDEC,PrePostAwd,PROT PER KTR\n")     

        fd.write(AP_05_09_CSV)

        fd.close()  

    ## Sustained Protests CSV
    with codecs.open (SP_05_09_CSV_path, 'w', 'utf-8') as fd:
        ## Write in the column headers you want to use
 
        fd.write("COMPANY NAME, DUNS  NO.,PROT REP'd,PROTS_CONT'd,SHORT CITE,CONTRACT NO.,SOLIC ISS DATE,PROT FILE DATE,PROT DEC DATE,ULT K AWD DATE,ORIGINATING AGENCY,PROTEST OUTC,WAS K RECOMP'd,ULT K REC'pnt,SB or DG?,PROC VAL,<<|>>,PROT PER K,ISS2PROT,PROT2DEC,PROT2ULTK,FY,PROTF2KDEC,PrePostAwd,PROT PER KTR\n")      

        fd.write(SP_05_09_CSV)

        fd.close()  

## 2010 - 2014 
    ## All Protests CSV
    with codecs.open (AP_10_14_CSV_path, 'w', 'utf-8') as fd:
        ## Write in the column headers you want to use
 
        fd.write("COMPANY NAME, DUNS  NO.,PROT REP'd,PROTS_CONT'd,SHORT CITE,CONTRACT NO.,SOLIC ISS DATE,PROT FILE DATE,PROT DEC DATE,ULT K AWD DATE,ORIGINATING AGENCY,PROTEST OUTC,WAS K RECOMP'd,ULT K REC'pnt,SB or DG?,PROC VAL,<<|>>,PROT PER K,ISS2PROT,PROT2DEC,PROT2ULTK,FY,PROTF2KDEC,PrePostAwd,PROT PER KTR\n")     

        fd.write(AP_10_14_CSV)

        fd.close()  

    ## Sustained Protests CSV
    with codecs.open (SP_10_14_CSV_path, 'w', 'utf-8') as fd:
        ## Write in the column headers you want to use
 
        fd.write("COMPANY NAME, DUNS  NO.,PROT REP'd,PROTS_CONT'd,SHORT CITE,CONTRACT NO.,SOLIC ISS DATE,PROT FILE DATE,PROT DEC DATE,ULT K AWD DATE,ORIGINATING AGENCY,PROTEST OUTC,WAS K RECOMP'd,ULT K REC'pnt,SB or DG?,PROC VAL,<<|>>,PROT PER K,ISS2PROT,PROT2DEC,PROT2ULTK,FY,PROTF2KDEC,PrePostAwd,PROT PER KTR\n")      

        fd.write(SP_10_14_CSV)

        fd.close()  

## 2015 - 2019 
    ## All Protests CSV
    with codecs.open (AP_15_19_CSV_path, 'w', 'utf-8') as fd:
        ## Write in the column headers you want to use
 
        fd.write("COMPANY NAME, DUNS  NO.,PROT REP'd,PROTS_CONT'd,SHORT CITE,CONTRACT NO.,SOLIC ISS DATE,PROT FILE DATE,PROT DEC DATE,ULT K AWD DATE,ORIGINATING AGENCY,PROTEST OUTC,WAS K RECOMP'd,ULT K REC'pnt,SB or DG?,PROC VAL,<<|>>,PROT PER K,ISS2PROT,PROT2DEC,PROT2ULTK,FY,PROTF2KDEC,PrePostAwd,PROT PER KTR\n")     

        fd.write(AP_15_19_CSV)

        fd.close()  

    ## Sustained Protests CSV
    with codecs.open (SP_15_19_CSV_path, 'w', 'utf-8') as fd:
        ## Write in the column headers you want to use
 
        fd.write("COMPANY NAME, DUNS  NO.,PROT REP'd,PROTS_CONT'd,SHORT CITE,CONTRACT NO.,SOLIC ISS DATE,PROT FILE DATE,PROT DEC DATE,ULT K AWD DATE,ORIGINATING AGENCY,PROTEST OUTC,WAS K RECOMP'd,ULT K REC'pnt,SB or DG?,PROC VAL,<<|>>,PROT PER K,ISS2PROT,PROT2DEC,PROT2ULTK,FY,PROTF2KDEC,PrePostAwd,PROT PER KTR\n")      

        fd.write(SP_15_19_CSV)

        fd.close()  

## 2020 - 2024 
    ## All Protests CSV
    with codecs.open (AP_20_24_CSV_path, 'w', 'utf-8') as fd:
        ## Write in the column headers you want to use
 
        fd.write("COMPANY NAME, DUNS  NO.,PROT REP'd,PROTS_CONT'd,SHORT CITE,CONTRACT NO.,SOLIC ISS DATE,PROT FILE DATE,PROT DEC DATE,ULT K AWD DATE,ORIGINATING AGENCY,PROTEST OUTC,WAS K RECOMP'd,ULT K REC'pnt,SB or DG?,PROC VAL,<<|>>,PROT PER K,ISS2PROT,PROT2DEC,PROT2ULTK,FY,PROTF2KDEC,PrePostAwd,PROT PER KTR\n")     

        fd.write(AP_20_24_CSV)

        fd.close()  

    ## Sustained Protests CSV
    with codecs.open (SP_20_24_CSV_path, 'w', 'utf-8') as fd:
        ## Write in the column headers you want to use
 
        fd.write("COMPANY NAME, DUNS  NO.,PROT REP'd,PROTS_CONT'd,SHORT CITE,CONTRACT NO.,SOLIC ISS DATE,PROT FILE DATE,PROT DEC DATE,ULT K AWD DATE,ORIGINATING AGENCY,PROTEST OUTC,WAS K RECOMP'd,ULT K REC'pnt,SB or DG?,PROC VAL,<<|>>,PROT PER K,ISS2PROT,PROT2DEC,PROT2ULTK,FY,PROTF2KDEC,PrePostAwd,PROT PER KTR\n")      

        fd.write(SP_20_24_CSV)

        fd.close()  



    ## Assholes
    ## -- These are all the protests which threw error codes
    ## -- over the course of the program. 

    with codecs.open (Ass_CSV_path, 'w', 'utf-8') as fd:
        ## Write in the column headers you want to use
 
        fd.write("ASSHOLES LIST\n")     

        fd.write(Ass_CSV)

        fd.close()  

############################################
##########################################
###
#####  END OF OUTPUT FILE MANAGEMENT
###
#######################################
#####################################


print("""
==================
  == ALL DONE ==
==================
""")

